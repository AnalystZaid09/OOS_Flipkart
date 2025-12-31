import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
import os
import traceback
import warnings
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(page_title="OOS Flipkart Sales Analysis", page_icon="📊", layout="wide")

# Title and description
st.title("📊 OOS Flipkart Sales Analysis Dashboard")
st.markdown("""
This application analyzes Flipkart sales data and calculates important metrics like:
- **DRR (Daily Run Rate)**: Average units sold per day
- **DOC (Days of Coverage)**: How many days current stock will last based on sales velocity
""")

st.markdown("---")

# Sidebar for inputs
with st.sidebar:
    st.header("⚙️ Configuration")
    st.markdown("### Enter Number of Days")
    st.info("📅 Enter the number of days covered in your sales report (e.g., 27 for Nov 1-27)")
    no_of_days = st.number_input(
        "Number of Days in Sales Period:",
        min_value=1,
        max_value=365,
        value=27,
        step=1,
        help="This is used to calculate Daily Run Rate (DRR)"
    )
    
    st.markdown("---")
    st.markdown("""
    ### 📈 DOC Color Legend:
    - 🔴 **Red (0-7 days)**: Critical - Immediate action needed
    - 🟠 **Orange (7-15 days)**: Low - Reorder soon
    - 🟢 **Green (15-30 days)**: Optimal - Good stock level
    - 🟡 **Yellow (30-45 days)**: Monitor sales
    - 🔵 **Sky Blue (45-60 days)**: High - Monitor closely
    - 🟤 **Brown (60-90 days)**: Excess - Stop ordering
    - ⬛ **Black (>90 days)**: Overstocked - Clearance needed
    """)

# Center column for file upload
col1, col2, col3 = st.columns([1, 2, 1])

with col2:
    st.markdown("### 📁 Upload Your Files")
    
    # File uploaders
    sales_file = st.file_uploader(
        "Upload Flipkart Sales Report (Excel/XLS/CSV)",
        type=['xlsx', 'xls','csv'],
        help="Upload the sales report downloaded from Flipkart"
    )
    
    inventory_file = st.file_uploader(
        "Upload Inventory Listing (Excel/XLS/CSV)",
        type=['xlsx', 'xls','csv'],
        help="Upload the inventory/listing report from Flipkart"
    )
    
    pm_file = st.file_uploader(
        "Upload Product Master (Excel/XLS/CSV)",
        type=['xlsx', 'xls','csv'],
        help="Upload the product master file with brand and manager details"
    )

st.markdown("---")

# 🔴🟠🟢 DOC styling for Streamlit table
def style_doc_column(s):
    styles = []
    for v in s:
        try:
            value = float(v)
        except (TypeError, ValueError):
            value = 0
        if 0 <= value < 7:
            styles.append('background-color: #FF0000; color: white;')
        elif 7 <= value < 15:
            styles.append('background-color: #FFA500; color: white;')  # orange
        elif 15 <= value < 30:
            styles.append('background-color: #008000; color: white;')  # green
        elif 30 <= value < 45:
            styles.append('background-color: #FFFF00; color: black;')  # yellow
        elif 45 <= value < 60:
            styles.append('background-color: #87CEEB; color: black;')  # sky blue
        elif 60 <= value < 90:
            styles.append('background-color: #8B4513; color: white;')  # brown
        else:  # >= 90
            styles.append('background-color: #000000; color: white;')  # black
    return styles

# ----- shared color fills for Excel DOC -----
DOC_FILLS = {
    "red": PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
    "orange": PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
    "green": PatternFill(start_color='008000', end_color='008000', fill_type='solid'),
    "yellow": PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
    "sky": PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid'),
    "brown": PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid'),
    "black": PatternFill(start_color='000000', end_color='000000', fill_type='solid'),
}
WHITE_FONT = Font(color="FFFFFF")
BLACK_FONT = Font(color="000000")

def apply_doc_color_to_column(ws, header_row_idx: int, col_name: str = "DOC"):
    """
    Loop DOC column cells in worksheet ws and apply same color logic as app.
    """
    # Find DOC column index
    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]
    doc_col_idx = None
    for idx, cell in enumerate(header_cells, start=1):
        if cell.value and str(cell.value).strip().lower() == col_name.lower():
            doc_col_idx = idx
            break
    if doc_col_idx is None:
        return

    max_row = ws.max_row
    for row in range(header_row_idx + 1, max_row + 1):
        cell = ws.cell(row=row, column=doc_col_idx)
        try:
            v = float(cell.value) if cell.value is not None else 0
        except (TypeError, ValueError):
            v = 0
        if 0 <= v < 7:
            cell.fill = DOC_FILLS["red"]
            cell.font = WHITE_FONT
        elif 7 <= v < 15:
            cell.fill = DOC_FILLS["orange"]
            cell.font = WHITE_FONT
        elif 15 <= v < 30:
            cell.fill = DOC_FILLS["green"]
            cell.font = WHITE_FONT
        elif 30 <= v < 45:
            cell.fill = DOC_FILLS["yellow"]
            cell.font = BLACK_FONT
        elif 45 <= v < 60:
            cell.fill = DOC_FILLS["sky"]
            cell.font = BLACK_FONT
        elif 60 <= v < 90:
            cell.fill = DOC_FILLS["brown"]
            cell.font = WHITE_FONT
        elif v >= 90:
            cell.fill = DOC_FILLS["black"]
            cell.font = WHITE_FONT

# ---------- Helper: safe Excel reader for xls/xlsx ----------
def read_excel_safely(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".xls"):
        # Requires xlrd<2.0.0 installed in the environment
        try:
            return pd.read_excel(uploaded_file, engine="xlrd")
        except Exception:
            # Fallback – may still fail if engine doesn't support .xls
            return pd.read_excel(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)

def read_file_safely(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    else:
        return read_excel_safely(uploaded_file)

# ---------- Helper 1: main formatted Excel ----------
def create_formatted_excel(df):
    output = BytesIO()
    
    # Write to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sales Analysis')
    
    # Load workbook for formatting
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Sales Analysis']
    
    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter on header row
    ws.auto_filter.ref = ws.dimensions

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_idx = col[0].column
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Apply DOC color formatting
    apply_doc_color_to_column(ws, header_row_idx=1, col_name="DOC")
    
    # Save to BytesIO
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final

# ---------- Helper 2: fill template (DataTable or first table or create new) ----------
def fill_template_and_get_bytes(template_path: str, df: pd.DataFrame, table_name: str = "DataTable") -> BytesIO:
    """
    Load an Excel template (xlsx/xlsm) and fill an Excel Table with `df`.

    Priority:
    1) Try table named `table_name` (e.g. DataTable)
    2) If not found but some table exists, use the first table
    3) If no tables at all, create a new sheet 'Data', write df, and create a new table `table_name`.

    Also applies DOC cell coloring on the filled sheet.
    """
    import re

    wb = load_workbook(template_path, keep_vba=True)
    table_sheet = None
    table_obj = None
    first_table_sheet = None
    first_table_obj = None

    # --- scan all sheets for tables ---
    for ws in wb.worksheets:
        tables = getattr(ws, "_tables", None)
        if not tables:
            continue

        if isinstance(tables, dict):
            iter_tables = list(tables.values())
        else:
            iter_tables = list(tables)

        for tbl in iter_tables:
            name = None
            try:
                name = getattr(tbl, "displayName", None) or getattr(tbl, "name", None)
            except Exception:
                pass

            if name is None and isinstance(tbl, str):
                name = tbl  # in case it's just a name string

            # remember the first table in the whole file
            if first_table_obj is None:
                first_table_obj = tbl
                first_table_sheet = ws

            # preferred: exact name match
            if name == table_name:
                table_sheet = ws
                table_obj = tbl
                break

        if table_obj is not None:
            break

    # CASE 1: some table exists but DataTable not found → use first table
    if table_obj is None and first_table_obj is not None:
        table_obj = first_table_obj
        table_sheet = first_table_sheet

    # CASE 2: no tables at all → create new sheet + table
    if table_obj is None or table_sheet is None:
        # choose or create a sheet named "Data"
        sheet_name = "Data"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # clear existing content
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)

        # write header + rows
        header = list(df.columns)
        ws.append(header)
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))

        # define table ref
        max_row = ws.max_row
        max_col = ws.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)

        # apply DOC coloring
        apply_doc_color_to_column(ws, header_row_idx=1, col_name="DOC")

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    # ----- helper: A1 -> (row, col) -----
    def cell_to_rowcol(cell_ref: str):
        m = re.match(r"([A-Z]+)(\d+)$", cell_ref)
        if not m:
            raise RuntimeError(f"Unexpected table ref format: {cell_ref}")
        col_letters, row = m.groups()
        col = 0
        for ch in col_letters:
            col = col * 26 + (ord(ch) - ord("A") + 1)
        return int(row), col

    # ----- clear old data & write df into existing table range -----
    ref = table_obj.ref             # e.g. "A1:H200"
    start_cell, end_cell = ref.split(":")
    start_row, start_col = cell_to_rowcol(start_cell)
    end_row, end_col = cell_to_rowcol(end_cell)

    # clear existing rows below header
    for r in range(start_row + 1, end_row + 1):
        for c in range(start_col, end_col + 1):
            table_sheet.cell(row=r, column=c).value = None

    # write header
    header = list(df.columns)
    for idx, col_name in enumerate(header):
        table_sheet.cell(row=start_row, column=start_col + idx, value=col_name)

    # write data rows
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=start_row + 1):
        for c_idx, v in enumerate(row, start=start_col):
            table_sheet.cell(row=r_idx, column=c_idx, value=v)

    # resize the table to new df
    new_end_row = start_row + len(df)
    new_end_col = start_col + len(header) - 1
    new_ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(new_end_col)}{new_end_row}"
    table_obj.ref = new_ref

    # apply DOC formatting to this sheet
    apply_doc_color_to_column(table_sheet, header_row_idx=start_row, col_name="DOC")

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ---------- Helper 3: fallback workbook with PivotSummary + Chart ----------
def create_pivot_fallback_workbook(df: pd.DataFrame, sheet_name: str, sort_ascending: bool = False) -> BytesIO:
    """
    Fallback workbook:
      - Data sheet with df (DOC colored)
      - DataTable
      - PivotSummary (Brand + Product Id → sum DOC & DRR, DOC colored)
      - ChartData + Chart
      - HowToPivot instructions

    sort_ascending:
      - True  → DOC ascending  (for OOS)
      - False → DOC descending (for Overstock)
    """
    working = df.copy()

    if "DOC" in working.columns:
        working["DOC"] = pd.to_numeric(working["DOC"], errors="coerce")
        working = working.sort_values(by="DOC", ascending=sort_ascending)

    if "DRR" in working.columns:
        working["DRR"] = pd.to_numeric(working["DRR"], errors="coerce")

    # Aggregate for summary
    if "Brand" in working.columns and "Product Id" in working.columns:
        agg = (
            working.groupby(["Brand", "Product Id"], dropna=False)[["DOC", "DRR"]]
            .sum()
            .reset_index()
        )
        agg["Brand_Parent"] = agg["Brand"].astype(str) + " | " + agg["Product Id"].astype(str)
    elif "Brand" in working.columns:
        agg = (
            working.groupby(["Brand"], dropna=False)[["DOC", "DRR"]]
            .sum()
            .reset_index()
        )
        agg["Brand_Parent"] = agg["Brand"].astype(str)
    else:
        agg = pd.DataFrame(columns=["Brand_Parent", "DOC", "DRR"])

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Data sheet
    for r in dataframe_to_rows(working, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto column width
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Color DOC in Data sheet
    apply_doc_color_to_column(ws, header_row_idx=1, col_name="DOC")

    # Add DataTable
    try:
        max_row = ws.max_row
        max_col = ws.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName="DataTable", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)
    except Exception:
        pass

    # PivotSummary sheet
    ws_pivot = wb.create_sheet("PivotSummary")
    if not agg.empty:
        for r in dataframe_to_rows(agg, index=False, header=True):
            ws_pivot.append(r)

        # Color DOC in PivotSummary
        apply_doc_color_to_column(ws_pivot, header_row_idx=1, col_name="DOC")

    # ChartData + Chart
    ws_chartdata = wb.create_sheet("ChartData")
    if not agg.empty:
        for r in dataframe_to_rows(agg[["Brand_Parent", "DOC", "DRR"]], index=False, header=True):
            ws_chartdata.append(r)
        if ws_chartdata.max_row > 1:
            chart = BarChart()
            cats = Reference(ws_chartdata, min_col=1, min_row=2, max_row=ws_chartdata.max_row)
            vals1 = Reference(ws_chartdata, min_col=2, min_row=2, max_row=ws_chartdata.max_row)
            vals2 = Reference(ws_chartdata, min_col=3, min_row=2, max_row=ws_chartdata.max_row)
            chart.add_data(vals1, titles_from_data=False)
            chart.add_data(vals2, titles_from_data=False)
            chart.set_categories(cats)
            chart.title = f"Sum DOC and DRR by Brand + Product ({sheet_name})"
            ws_chart = wb.create_sheet("Chart")
            ws_chart.add_chart(chart, "A1")

    # HowToPivot sheet
    ws_how = wb.create_sheet("HowToPivot")
    ws_how.append([f"How to create PivotTable + PivotChart ({sheet_name})"])
    ws_how.append([])
    ws_how.append(["1) Insert → PivotTable using table 'DataTable' on sheet " + sheet_name])
    ws_how.append(["2) Put 'Brand' in Rows, then 'Product Id' under it."])
    ws_how.append(["3) Put 'DOC' and 'DRR' into Values as Sum."])
    ws_how.append(["4) Insert → PivotChart from PivotTable."])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

tab_business, tab_inventory = st.tabs(["📊 Business Report", "📦 Inventory Report"])

# ------------------------- PROCESS DATA BUTTON -------------------------
if sales_file and inventory_file and pm_file:
    with tab_business:
        if st.button("🚀 Process Data", type="primary", use_container_width=True):
            with st.spinner("Processing your data..."):
                try:
                    # Read files safely
                    F_Sales = read_file_safely(sales_file)
                    Inventory = read_file_safely(inventory_file)
                    PM = read_file_safely(pm_file)
                    
                    # Basic column checks for Sales
                    required_sales_cols = {"Product Id", "Final Sale Units"}
                    missing_sales = required_sales_cols - set(F_Sales.columns)
                    if missing_sales:
                        st.error(f"Sales file is missing these columns: {', '.join(missing_sales)}")
                        st.stop()

                    # Ensure Final Sale Units numeric & clean negatives
                    if isinstance(F_Sales, pd.DataFrame) and "Final Sale Units" in F_Sales.columns:
                        F_Sales["Final Sale Units"] = pd.to_numeric(
                            F_Sales["Final Sale Units"], errors="coerce"
                        ).fillna(0)
                        F_Sales.loc[F_Sales["Final Sale Units"] < 0, "Final Sale Units"] = 0

                    # Store raw total units BEFORE any merges/grouping (to match original file)
                    raw_total_units = F_Sales["Final Sale Units"].sum()
                    st.session_state["raw_total_units"] = float(raw_total_units)

                    # ------- GROUP DUPLICATE PRODUCT ID & SUM Final Sale Units -------
                    if isinstance(F_Sales, pd.DataFrame) and "Product Id" in F_Sales.columns:
                        dup_count_pid = int(F_Sales.duplicated(subset=["Product Id"]).sum())
                        if dup_count_pid > 0:
                            st.info("🔁 Duplicate Product Id found — combining rows and summing Final Sale Units")

                            agg_dict = {}
                            for col in F_Sales.columns:
                                if col == "Final Sale Units":
                                    agg_dict[col] = "sum"
                                elif col != "Product Id":
                                    # keep first value for all other columns
                                    agg_dict[col] = "first"

                            st.success(f"🧮 Aggregated {dup_count_pid} duplicate Product Id rows (Final Sale Units summed).")
                        else:
                            st.info("ℹ️ No duplicate Product Id values found to aggregate.")
                    else:
                        if isinstance(F_Sales, pd.DataFrame):
                            st.warning("⚠️ 'Product Id' column not found — Product Id aggregation skipped.")

                    # Merge with PM (Product Master) safely: de-duplicate FNS
                    PM_cols_needed = {"FNS", "Brand Manager", "Brand"}
                    missing_pm = PM_cols_needed - set(PM.columns)
                    if missing_pm:
                        st.error(f"Product Master file is missing these columns: {', '.join(missing_pm)}")
                        st.stop()

                    pm_small = PM[['FNS', 'Brand Manager', 'Brand']].copy()
                    pm_small = pm_small.drop_duplicates(subset=['FNS'], keep='first')

                    F_Sales = F_Sales.merge(
                        pm_small,
                        left_on='Product Id',
                        right_on='FNS',
                        how='left'
                    )
                    F_Sales = F_Sales.drop(columns=['FNS'], errors='ignore')
                    
                    # ---------------- PURCHASE MASTER COLUMNS (Vendor SKU, Product Name, CP) ----------------

                    required_pm_cols = {
                        "FNS",
                        "Vendor SKU Codes",
                        "Product Name",
                        "CP"
                    }

                    missing_pm_cols = required_pm_cols - set(PM.columns)
                    if missing_pm_cols:
                        st.error(f"Purchase Master file missing columns: {', '.join(missing_pm_cols)}")
                        st.stop()

                    pm_purchase = PM[
                        ["FNS", "Vendor SKU Codes", "Product Name", "CP"]
                    ].copy()

                    # De-duplicate on Product Id
                    pm_purchase = pm_purchase.drop_duplicates(
                        subset=["FNS"],
                        keep="first"
                    )

                    # Merge into Business Report
                    F_Sales = F_Sales.merge(
                        pm_purchase,
                        left_on="Product Id",
                        right_on="FNS",
                        how="left"
                    )

                    # Rename and reorganize columns
                    if 'Brand_y' in F_Sales.columns:
                        F_Sales.rename(columns={'Brand_y': 'Brand'}, inplace=True)
                        if 'Brand_x' in F_Sales.columns:
                            F_Sales.drop(columns=['Brand_x'], inplace=True)
                    
                    # Reorder columns
                    cols = F_Sales.columns.tolist()
                    if 'Brand Manager' in cols and 'Brand' in cols and 'SKU ID' in cols:
                        bm = cols.pop(cols.index('Brand Manager'))
                        br = cols.pop(cols.index('Brand'))
                        sku_pos = cols.index('SKU ID')
                        cols.insert(sku_pos, br)
                        cols.insert(sku_pos, bm)
                        F_Sales = F_Sales[cols]
                        
                    # 🔥 YAHAN FILTER LAGAO (merge ke baad)
                    if "Final Sale Units" in F_Sales.columns:
                        F_Sales["Final Sale Units"] = pd.to_numeric(F_Sales["Final Sale Units"], errors="coerce")
                        F_Sales = F_Sales[F_Sales["Final Sale Units"] > 0]
                    
                    if "Final Sale Amount" in F_Sales.columns:
                        F_Sales["Final Sale Amount"] = pd.to_numeric(F_Sales["Final Sale Amount"], errors="coerce")
                        F_Sales = F_Sales[F_Sales["Final Sale Amount"] > 0]

                    # Calculate DRR using cleaned Final Sale Units
                    if "Final Sale Units" in F_Sales.columns:
                        F_Sales["DRR"] = (F_Sales["Final Sale Units"] / no_of_days).round(2)
                    else:
                        F_Sales["DRR"] = 0

                    # --- Clean Inventory columns and ensure correct mapping (de-duplicate key) ---
                    inv = Inventory.copy()
                    inv.columns = inv.columns.str.strip()

                    required_inv_cols = {"Flipkart Serial Number", "System Stock count"}
                    missing_inv = required_inv_cols - set(inv.columns)
                    if missing_inv:
                        st.error(f"Inventory file is missing these columns: {', '.join(missing_inv)}")
                        st.stop()

                    inv_small = inv[["Flipkart Serial Number", "System Stock count"]].drop_duplicates(
                        subset=["Flipkart Serial Number"],
                        keep="first"
                    )

                    # Merge Sales with Inventory
                    F_Sales = F_Sales.merge(
                        inv_small,
                        left_on="Product Id",
                        right_on="Flipkart Serial Number",
                        how="left"
                    )

                    F_Sales.rename(columns={"System Stock count": "Flipkart Stock"}, inplace=True)
                    F_Sales.drop(columns=["Flipkart Serial Number"], errors="ignore", inplace=True)
                    
                    # Calculate DOC
                    F_Sales["Flipkart Stock"] = pd.to_numeric(F_Sales["Flipkart Stock"], errors="coerce")
                    F_Sales["DRR"] = pd.to_numeric(F_Sales["DRR"], errors="coerce")
                    
                    F_Sales["DOC"] = np.where(
                        F_Sales["DRR"] > 0,
                        F_Sales["Flipkart Stock"] / F_Sales["DRR"],
                        np.nan
                    )
                    F_Sales["DOC"] = F_Sales["DOC"].round(2)
                    F_Sales["DOC"] = F_Sales["DOC"].fillna(0)
                    
                    cols = F_Sales.columns.tolist()

                    if "Product Id" in cols:
                        insert_at = cols.index("Product Id") + 1
                        for c in ["Vendor SKU Codes", "Product Name"]:
                            if c in cols:
                                cols.insert(insert_at, cols.pop(cols.index(c)))
                                insert_at += 1

                    F_Sales = F_Sales[cols]

                    # MAIN processed full Excel (no filters)
                    excel_data_bytes = create_formatted_excel(F_Sales).getvalue()

                    # OOS: Flipkart Stock == 0, DOC ascending
                    oos_df = F_Sales.copy()
                    oos_df["Flipkart Stock"] = pd.to_numeric(oos_df["Flipkart Stock"], errors="coerce").fillna(0)
                    oos_df = oos_df[oos_df["Flipkart Stock"] == 0].copy()
                    if "DOC" in oos_df.columns:
                        oos_df["DOC"] = pd.to_numeric(oos_df["DOC"], errors="coerce")
                        oos_df = oos_df.sort_values(by="DOC", ascending=True).reset_index(drop=True)

                    # Overstock: DOC >= 90, DOC descending
                    over_df = F_Sales.copy()
                    over_df["DOC"] = pd.to_numeric(over_df["DOC"], errors="coerce")
                    over_df = over_df[over_df["DOC"] >= 90].copy()
                    over_df = over_df.sort_values(by="DOC", ascending=False).reset_index(drop=True)

                    base_dir = os.path.dirname(__file__)
                    tmpl_xlsm = os.path.join(base_dir, "flipkart_pivot_template.xlsm")
                    tmpl_xlsx = os.path.join(base_dir, "flipkart_pivot_template.xlsx")
                    template_path = tmpl_xlsm if os.path.exists(tmpl_xlsm) else (tmpl_xlsx if os.path.exists(tmpl_xlsx) else None)

                    # ---------- OOS export ----------
                    oos_bytes = None
                    oos_ext = ".xlsx"
                    oos_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                    if template_path:
                        try:
                            buf_oos = fill_template_and_get_bytes(template_path, oos_df, table_name="DataTable")
                            oos_bytes = buf_oos.getvalue()
                            if template_path.lower().endswith(".xlsm"):
                                oos_ext = ".xlsm"
                                oos_mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
                            st.success("✅ OOS: Used pivot template — your VBA macro will build PivotTable & PivotChart on open.")
                        except Exception:
                            st.warning("⚠️ OOS: Template fill failed, using fallback workbook.")
                            with st.expander("Show OOS error details"):
                                st.code(traceback.format_exc())

                    if oos_bytes is None:
                        fb_oos = create_pivot_fallback_workbook(oos_df, sheet_name="OOS", sort_ascending=True)
                        oos_bytes = fb_oos.getvalue()
                        oos_ext = ".xlsx"
                        oos_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        st.info("ℹ️ OOS fallback workbook generated (DataTable + PivotSummary + ChartData + HowToPivot).")

                    # ---------- Overstock export ----------
                    over_bytes = None
                    over_ext = ".xlsx"
                    over_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

                    if template_path:
                        try:
                            buf_over = fill_template_and_get_bytes(template_path, over_df, table_name="DataTable")
                            over_bytes = buf_over.getvalue()
                            if template_path.lower().endswith(".xlsm"):
                                over_ext = ".xlsm"
                                over_mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
                            st.success("✅ Overstock: Used pivot template — your VBA macro will build PivotTable & PivotChart on open.")
                        except Exception:
                            st.warning("⚠️ Overstock: Template fill failed, using fallback workbook.")
                            with st.expander("Show Overstock error details"):
                                st.code(traceback.format_exc())

                    if over_bytes is None:
                        fb_over = create_pivot_fallback_workbook(over_df, sheet_name="Overstock", sort_ascending=False)
                        over_bytes = fb_over.getvalue()
                        over_ext = ".xlsx"
                        over_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        st.info("ℹ️ Overstock fallback workbook generated (DataTable + PivotSummary + ChartData + HowToPivot).")

                    # Save into session_state so download buttons can be used anytime
                    st.session_state["F_Sales_df"] = F_Sales
                    st.session_state["excel_data_main"] = excel_data_bytes
                    st.session_state["oos_bytes"] = oos_bytes
                    st.session_state["oos_ext"] = oos_ext
                    st.session_state["oos_mime"] = oos_mime
                    st.session_state["over_bytes"] = over_bytes
                    st.session_state["over_ext"] = over_ext
                    st.session_state["over_mime"] = over_mime

                    st.success("✅ Data processed and files prepared! Scroll down to view and download.")
                
                except Exception as e:
                    st.error(f"❌ Error processing files: {str(e)}")
                    st.info("Please ensure all files are in the correct format and contain the required columns.")
                    
                # ---------- UNIQUE PRODUCT ID REPORT (CUSTOM COLUMNS + AGG RULES) ----------
                st.markdown("---")
                st.subheader("📊 Unique Product ID Report (Summed as per rules)")

                # Select only required columns if exist
                cols_needed = ["Product Id", "Vendor SKU Codes", "Product Name", "Brand Manager", "Brand", 
                            "SKU ID", "Final Sale Units", "Final Sale Amount", "FNS", "CP", "DRR",
                            "Flipkart Stock", "DOC"]

                unique_df = F_Sales[[c for c in cols_needed if c in F_Sales.columns]].copy()

                # Ensure numeric conversion
                sum_cols = ["Final Sale Units", "Final Sale Amount", "CP", "DRR", "Flipkart Stock", "DOC"]
                for c in sum_cols:
                    if c in unique_df.columns:
                        unique_df[c] = pd.to_numeric(unique_df[c], errors="coerce").fillna(0)

                # Aggregation logic
                agg_rules = {}
                for col in unique_df.columns:
                    if col == "Product Id":
                        continue
                    elif col in ["Vendor SKU Codes", "Product Name", "Brand Manager", "Brand", "SKU ID"]:
                        agg_rules[col] = "first"   # take first occurrence
                    elif col in sum_cols:
                        agg_rules[col] = "sum"     # sum numeric columns
                    else:
                        agg_rules[col] = "first"   # default safe fallback

                unique_df = unique_df.groupby("Product Id", as_index=False).agg(agg_rules)

                # Display in Streamlit
                st.dataframe(unique_df, height=350, use_container_width=True)

                # Download button
                unique_excel = BytesIO()
                with pd.ExcelWriter(unique_excel, engine="openpyxl") as writer:
                    unique_df.to_excel(writer, index=False, sheet_name="Unique_PID")
                unique_excel.seek(0)

                st.download_button(
                    label="📥 Download Unique Product ID Report",
                    data=unique_excel,
                    file_name="Flipkart_Sales_Unique_ProductID_Summed.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                st.success("✅ Unique Product ID report ready!")

        # ------------------------- VIEW + DOWNLOAD SECTION -------------------------
        if "F_Sales_df" in st.session_state:
            F_Sales = st.session_state["F_Sales_df"]

            st.markdown("---")
            st.header("📈 Processed Results")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Products", len(F_Sales))
            with col2:
                if "raw_total_units" in st.session_state:
                    st.metric("Total Final Sale Units", int(st.session_state["raw_total_units"]))
                elif "Final Sale Units" in F_Sales.columns:
                    st.metric("Total Final Sale Units", int(F_Sales["Final Sale Units"].sum()))
                else:
                    st.metric("Total Final Sale Units", "N/A")
            with col3:
                if "Final Sale Amount" in F_Sales.columns:
                    st.metric("Total GMV", f"₹{F_Sales['Final Sale Amount'].sum():,.0f}")
                else:
                    st.metric("Total GMV", "N/A")
            with col4:
                if "DOC" in F_Sales.columns:
                    st.metric("Avg DOC", f"{F_Sales['DOC'].mean():.1f} days")
                else:
                    st.metric("Avg DOC", "N/A")
            
            st.markdown("### 📊 Processed Data Preview")
            if "DOC" in F_Sales.columns:
                styled_df = F_Sales.style.apply(style_doc_column, subset=['DOC'])
                st.dataframe(styled_df, height=400, use_container_width=True)
            else:
                st.dataframe(F_Sales, height=400, use_container_width=True)

            st.markdown("---")
            st.header("💾 Download Files")

            colA, colB, colC = st.columns(3)

            with colA:
                st.markdown("**Main Processed File**")
                st.download_button(
                    label="📥 Download Full Analysis Excel",
                    data=st.session_state["excel_data_main"],
                    file_name="Flipkart_Sales_Analysis_Formatted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="main_download",
                    use_container_width=True
                )

            with colB:
                st.markdown("**OOS (Flipkart Stock = 0, DOC ↑)**")
                st.download_button(
                    label=f"📥 Download OOS with Pivot ({st.session_state['oos_ext']})",
                    data=st.session_state["oos_bytes"],
                    file_name=f"Flipkart_OOS_with_Pivot{st.session_state['oos_ext']}",
                    mime=st.session_state["oos_mime"],
                    key="oos_download_btn",
                    use_container_width=True
                )

            with colC:
                st.markdown("**Overstock (DOC ≥ 90, DOC ↓)**")
                st.download_button(
                    label=f"📥 Download Overstock with Pivot ({st.session_state['over_ext']})",
                    data=st.session_state["over_bytes"],
                    file_name=f"Flipkart_Overstock_with_Pivot{st.session_state['over_ext']}",
                    mime=st.session_state["over_mime"],
                    key="over_download_btn",
                    use_container_width=True
                )

            st.markdown("---")
            st.markdown("### 📈 Stock Status Insights")
            col1, col2 = st.columns(2)
            
            if "DOC" in F_Sales.columns:
                with col1:
                    critical_stock = len(F_Sales[F_Sales['DOC'] < 7])
                    low_stock = len(F_Sales[(F_Sales['DOC'] >= 7) & (F_Sales['DOC'] < 15)])
                    optimal_stock = len(F_Sales[(F_Sales['DOC'] >= 15) & (F_Sales['DOC'] < 30)])
                    
                    st.markdown(f"""
                    **Stock Alerts:**
                    - 🔴 Critical (0-7 days): **{critical_stock} products**
                    - 🟠 Low (7-15 days): **{low_stock} products**
                    - 🟢 Optimal (15-30 days): **{optimal_stock} products**
                    """)
                
                with col2:
                    high_stock = len(F_Sales[(F_Sales['DOC'] >= 30) & (F_Sales['DOC'] < 45)])
                    very_high_stock = len(F_Sales[(F_Sales['DOC'] >= 45) & (F_Sales['DOC'] < 60)])
                    excess_stock = len(F_Sales[F_Sales['DOC'] >= 60])
                    
                    st.markdown(f"""
                    **Excess Stock:**
                    - 🟡 Monitor (30-45 days): **{high_stock} products**
                    - 🔵 High (45-60 days): **{very_high_stock} products**
                    - 🟤 Excess (60+ days): **{excess_stock} products**
                    """)
            else:
                st.info("DOC column not found, cannot compute stock buckets.")

    with tab_inventory:
        st.header("📦 Inventory Report")

        # -------------------------------------------------
        # STEP 1: INVENTORY HEADER CLEANING
        # -------------------------------------------------
        if "Inventory" in locals() and isinstance(Inventory, pd.DataFrame):

            inv = Inventory.copy()

            # Drop fully empty rows
            inv = inv.dropna(how="all")

            # First row as header
            inv.columns = inv.iloc[0]
            inv = inv.iloc[1:]

            # Clean column names
            inv.columns = inv.columns.astype(str).str.strip()

            # Reset index
            inv = inv.reset_index(drop=True)

            st.subheader("📄 Cleaned Inventory Preview")
            st.caption(f"Rows: {len(inv)} | Columns: {len(inv.columns)}")

            with st.expander("🔍 Inventory Columns"):
                st.write(list(inv.columns))

            st.dataframe(inv, use_container_width=True, height=300)

            # Save cleaned inventory
            st.session_state["Inventory_df"] = inv

            st.success("✅ Inventory header fixed and data cleaned.")

        else:
            st.warning("⚠️ Inventory data not available.")
            st.stop()

        # -------------------------------------------------
        # STEP 2: INVENTORY BUSINESS LOGIC
        # -------------------------------------------------
        if (
            "Inventory_df" in st.session_state
            and "F_Sales_df" in st.session_state
        ):

            inv = st.session_state["Inventory_df"].copy()
            sales = st.session_state["F_Sales_df"].copy()

            # ---------------- BUSINESS PIVOT (VLOOKUP LOGIC) ----------------
            business_pivot = (
                sales
                .groupby("Product Id", dropna=False)["Final Sale Units"]
                .sum()
                .reset_index()
                .rename(columns={"Final Sale Units": "Business Sales Qty"})
            )

            # ---------------- PURCHASE MASTER LOOKUP ----------------
            pm_inv = PM[
                ["FNS", "Brand", "Brand Manager", "Vendor SKU Codes", "CP"]
            ].copy()

            pm_inv = pm_inv.drop_duplicates(subset=["FNS"], keep="first")

            # ---------------- MERGE: INVENTORY + BUSINESS SALES ----------------
            inv = inv.merge(
                business_pivot,
                left_on="Flipkart's Identifier of the product",
                right_on="Product Id",
                how="left"
            )
            inv.drop(columns=["Product Id"], errors="ignore", inplace=True)

            # ---------------- MERGE: INVENTORY + PURCHASE MASTER ----------------
            inv = inv.merge(
                pm_inv,
                left_on="Flipkart's Identifier of the product",
                right_on="FNS",
                how="left"
            )
            inv.drop(columns=["FNS"], errors="ignore", inplace=True)

            # ---------------- CLEAN NUMERIC FIELDS ----------------
            inv["Current Stock Count"] = pd.to_numeric(
                inv["Current stock count for your product"], errors="coerce"
            ).fillna(0)

            inv["Business Sales Qty"] = pd.to_numeric(
                inv["Business Sales Qty"], errors="coerce"
            ).fillna(0)

            inv["CP"] = pd.to_numeric(inv["CP"], errors="coerce").fillna(0)

            # ---------------- CALCULATIONS ----------------
            inv["As per Qty"] = (inv["CP"] * inv["Current Stock Count"]).round(2)

            inv["DRR"] = np.where(
                inv["Business Sales Qty"] > 0,
                (inv["Business Sales Qty"] / no_of_days).round(2),
                0
            )

            inv["DOC"] = np.where(
                inv["DRR"] > 0,
                (inv["Current Stock Count"] / inv["DRR"]).round(2),
                0
            )

            # ---------------- FINAL COLUMN ORDER ----------------
            preferred_cols = [
                "Flipkart's Identifier of the product",
                "Vendor SKU Codes",
                "Brand",
                "Brand Manager",
                "CP",
                "Current Stock Count",
                "As per Qty",
                "Business Sales Qty",
                "DRR",
                "DOC"
            ]

            final_cols = (
                [c for c in preferred_cols if c in inv.columns] +
                [c for c in inv.columns if c not in preferred_cols]
            )

            inv = inv[final_cols]

            # ---------------- DISPLAY FINAL INVENTORY REPORT ----------------
            st.subheader("📊 Inventory Analysis (Final)")
            st.dataframe(inv, use_container_width=True, height=350)

            st.success("✅ Inventory report generated successfully.")

            st.session_state["Inventory_Report_Final"] = inv
            
            # ---------------- INVENTORY DOWNLOAD ----------------
            st.markdown("---")
            st.subheader("💾 Download Inventory Report")

            inventory_excel = BytesIO()

            with pd.ExcelWriter(inventory_excel, engine="openpyxl") as writer:
                inv.to_excel(writer, index=False, sheet_name="Inventory_Report")

            inventory_excel.seek(0)

            st.download_button(
                label="📥 Download Inventory Report (Excel)",
                data=inventory_excel,
                file_name="Flipkart_Inventory_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            # -------------------------------------------------
            # STEP 3: ZERO BUSINESS SALES QTY (NO SALES ITEMS)
            # -------------------------------------------------
            st.markdown("---")
            st.subheader("🚫 Inventory with ZERO Business Sales Qty")

            # Filter only Business Sales Qty == 0
            zero_sales_inv = inv[
                (inv["Business Sales Qty"].fillna(0) == 0)
            ].copy()

            st.caption(f"Products with zero sales: {len(zero_sales_inv)}")

            if zero_sales_inv.empty:
                st.info("ℹ️ No products found with Business Sales Qty = 0")
            else:
                st.dataframe(
                    zero_sales_inv,
                    use_container_width=True,
                    height=350
                )

                # ---------------- DOWNLOAD: ZERO SALES INVENTORY ----------------
                zero_sales_excel = BytesIO()

                with pd.ExcelWriter(zero_sales_excel, engine="openpyxl") as writer:
                    zero_sales_inv.to_excel(
                        writer,
                        index=False,
                        sheet_name="Zero_Business_Sales"
                    )

                zero_sales_excel.seek(0)

                st.download_button(
                    label="📥 Download ZERO Sales Inventory (Excel)",
                    data=zero_sales_excel,
                    file_name="Flipkart_Inventory_Zero_Business_Sales.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


        else:
            st.warning("⚠️ Please process Business Report first.")
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: gray;'>
        <p>Flipkart Sales Analysis Dashboard | Built with Streamlit</p>
    </div>
    """, unsafe_allow_html=True)
